import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from pathlib import Path
from typing import Dict, List
from datetime import datetime
from loguru import logger

class ExcelManager:
    def __init__(self, config: Dict):
        self.config = config
        self.output_path = Path(config['export']['excel_path'])
        self.output_path.mkdir(parents=True, exist_ok=True)
        
    def create_kpi_workbook(self, all_company_data: Dict[str, pd.DataFrame], 
                           company_mappings: pd.DataFrame) -> Path:
        """
        Create Excel workbook with all company KPI data
        
        Args:
            all_company_data: Dict mapping CSIN to DataFrame
            company_mappings: DataFrame with ticker/CSIN mappings
        
        Returns:
            Path to saved Excel file
        """
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create summary sheet
        self._create_summary_sheet(wb, all_company_data, company_mappings)
        
        # Create sheet for each company
        for company_id, df in all_company_data.items():
            ticker = company_mappings[company_mappings['company_id'] == company_id]['search_ticker'].iloc[0]
            self._create_company_sheet(wb, ticker, df)
        
        # Create config and documentation sheets
        self._create_config_sheet(wb, company_mappings)
        self._create_readme_sheet(wb)
        
        # Save workbook
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"KPI_Dashboard_{timestamp}.xlsx"
        filepath = self.output_path / filename
        
        wb.save(filepath)
        logger.info(f"Saved Excel workbook to {filepath}")
        
        return filepath
    
    def _create_company_sheet(self, wb: Workbook, ticker: str, df: pd.DataFrame):
        """Create formatted sheet for a single company"""
        ws = wb.create_sheet(ticker)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        # Row 1: Period type headers
        ws['A1'] = 'KPI Code'
        ws['B1'] = 'KPI Description'
        ws['C1'] = 'Units'
        
        col_idx = 4
        # Annual headers
        ws.cell(row=1, column=col_idx, value='Annual Data').font = header_font
        ws.cell(row=1, column=col_idx, value='Annual Data').fill = header_fill
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+4)
        
        # Quarterly headers
        col_idx = 9
        ws.cell(row=1, column=col_idx, value='Quarterly Data').font = header_font
        ws.cell(row=1, column=col_idx, value='Quarterly Data').fill = header_fill
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+11)
        
        # Row 2: Actual period names
        ws['A2'] = 'KPI Code'
        ws['B2'] = 'KPI Description'
        ws['C2'] = 'Units'
        
        col_idx = 4
        for col in df.columns:
            ws.cell(row=2, column=col_idx, value=col)
            ws.cell(row=2, column=col_idx).font = header_font
            ws.cell(row=2, column=col_idx).fill = header_fill
            col_idx += 1
        
        # Write data
        row_idx = 3
        for idx, row in df.iterrows():
            # Handle multi-index (time_series, description, units) and growth type
            if isinstance(idx, tuple):
                if len(idx) == 4 and pd.isna(idx[3]):  # Base row (4th element is nan)
                    ws.cell(row=row_idx, column=1, value=idx[0])  # KPI Code
                    ws.cell(row=row_idx, column=2, value=idx[1])  # Description
                    ws.cell(row=row_idx, column=3, value=idx[2])  # Units
                elif len(idx) == 4 and pd.notna(idx[3]):  # Growth row
                    ws.cell(row=row_idx, column=1, value=f"{idx[0]} - {idx[3]}")
                    ws.cell(row=row_idx, column=2, value=f"{idx[1]} - {idx[3]}")
                    ws.cell(row=row_idx, column=3, value="%")
            
            # Write values
            col_idx = 4
            for value in row:
                if pd.notna(value):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Apply number formatting
                    if 'qoq growth' in str(idx) or 'yoy growth' in str(idx):
                        cell.number_format = '0%'  # Show as percentage
                        cell.value = value / 100  # Convert to Excel percentage
                    elif isinstance(idx, tuple) and len(idx) >= 3 and (idx[2] == 'USD' or idx[2] == 'Millions'):
                        # Format as thousands with comma separator
                        cell.number_format = '#,##0'
                    elif isinstance(idx, tuple) and len(idx) >= 3 and idx[2] == 'Percentage':
                        cell.number_format = '0.0%'
                        cell.value = value / 100  # Convert to Excel percentage
                    elif isinstance(idx, tuple) and len(idx) >= 3 and idx[2] == 'Per Share':
                        cell.number_format = '$0.00'
                    
                    # Apply conditional formatting for growth rates
                    if 'qoq growth' in str(idx) or 'yoy growth' in str(idx):
                        if value > 0:
                            cell.font = Font(color="006100")
                        elif value < 0:
                            cell.font = Font(color="9C0006")
                
                col_idx += 1
            
            row_idx += 1
        
        # Apply borders and column widths
        for row in ws.iter_rows(min_row=1, max_row=row_idx-1, min_col=1, max_col=col_idx-1):
            for cell in row:
                cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        for col in range(4, col_idx):
            ws.column_dimensions[get_column_letter(col)].width = 12
        
        # Freeze panes
        ws.freeze_panes = 'D3'
    
    def _create_summary_sheet(self, wb: Workbook, all_company_data: Dict, 
                            company_mappings: pd.DataFrame):
        """Create summary dashboard"""
        ws = wb.create_sheet('Summary', 0)
        
        # Title
        ws['A1'] = 'KPI Dashboard Summary'
        ws['A1'].font = Font(size=16, bold=True)
        
        # Refresh info
        ws['A3'] = 'Last Refresh:'
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S %Z')
        
        # Company summary
        ws['A5'] = 'Companies Included:'
        row_idx = 6
        for company_id in all_company_data.keys():
            company_info = company_mappings[company_mappings['company_id'] == company_id].iloc[0]
            ws[f'A{row_idx}'] = company_info['search_ticker']
            ws[f'B{row_idx}'] = company_info['name']
            ws[f'C{row_idx}'] = company_info['sector']
            row_idx += 1
    
    def _create_config_sheet(self, wb: Workbook, company_mappings: pd.DataFrame):
        """Create configuration reference sheet"""
        ws = wb.create_sheet('Config')
        
        # Write company mappings
        ws['A1'] = 'Company Configuration'
        ws['A1'].font = Font(bold=True, size=14)
        
        # Headers
        headers = ['Ticker', 'CSIN', 'Company Name', 'Sector']
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header).font = Font(bold=True)
        
        # Data
        for idx, row in company_mappings.iterrows():
            ws.cell(row=idx+4, column=1, value=row['search_ticker'])
            ws.cell(row=idx+4, column=2, value=row['company_id'])
            ws.cell(row=idx+4, column=3, value=row['name'])
            ws.cell(row=idx+4, column=4, value=row['sector'])
    
    def _create_readme_sheet(self, wb: Workbook):
        """Create documentation sheet"""
        ws = wb.create_sheet('README')
        
        readme_content = [
            ('KPI Dashboard Documentation', 16, True),
            ('', 12, False),
            ('Overview', 14, True),
            ('This workbook contains financial KPI data retrieved from the Canalyst/Tegus API.', 12, False),
            ('', 12, False),
            ('Sheet Descriptions:', 14, True),
            ('- Summary: Overview of included companies and refresh status', 12, False),
            ('- [Ticker] sheets: Individual company KPI data with growth calculations', 12, False),
            ('- Config: Company ticker to CSIN mappings', 12, False),
            ('', 12, False),
            ('Data Layout:', 14, True),
            ('- Columns D-H: Annual data (FY24 to FY20)', 12, False),
            ('- Columns I-T: Quarterly data (Q1-25 to Q2-22)', 12, False),
            ('- Each KPI has 3 rows: Value, QoQ %, YoY %', 12, False),
            ('', 12, False),
            ('Refresh Instructions:', 14, True),
            ('Run: python main.py refresh', 12, False),
        ]
        
        for idx, (text, size, bold) in enumerate(readme_content, 1):
            ws[f'A{idx}'] = text
            ws[f'A{idx}'].font = Font(size=size, bold=bold)